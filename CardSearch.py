import tkinter as tk
import tkinter.font
import requests
from MagicCard import MagicCard
import io
from PIL import Image, ImageTk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os


class CardSearch:

    def __init__(self, username, listName):
        self.fileName = 'lists/' + username + "/" + listName
        self.HEIGHT = 1200
        self.WIDTH = 1697
        if os.path.exists(self.fileName):
            self.wb = load_workbook(filename=self.fileName)
            self.ws1 = self.wb.active
        else:
            self.fileName = self.fileName + ".xlsx"
            self.wb = Workbook()
            self.ws1 = self.wb.active
            values = ['Card Name', 'Set', 'Price(USD)', 'Rarity']
            row = 1
            column=1
            for value in values:
                self.ws1.cell(column=column, row=row, value=str(value))
                column+=1

        self.root = tk.Toplevel()
        font = tk.font.Font(family='Beleren Small Caps', size=25)

        canvas = tk.Canvas(self.root, height=self.HEIGHT, width=self.WIDTH)
        canvas.pack()

        self.background_image = tk.PhotoImage(file="background.png")
        self.background_label = tk.Label(self.root, image=self.background_image)
        self.background_label.place(relwidth=1, relheight=1)

        self.frame = tk.Frame(self.root, bg='#e7e6e7')
        self.frame.place(relwidth=0.75, relheight=.035, relx=0.5, rely=0.025, anchor='n')

        self.entry = tk.Entry(self.frame, bg='#e7e6e7', font=font)
        self.entry.insert(0, "Enter A Card Name")
        self.entry.place(relwidth=0.7, relheight=0.98)

        self.button = tk.Button(self.frame, bg='#e7e6e7', text="Search Cards", font=font,
                                command=lambda: self.get_cards(self.entry.get()))
        self.button.place(relx=.7, relheight=1, relwidth=.30)

        self.lower_frame = tk.Frame(self.root)
        self.lower_frame.place(rely=0.2, relx=0.5, width=1460, height=816, anchor='n')

        self.frame_background = tk.PhotoImage(file='backgroundFrame.png')
        self.frame_label = tk.Label(self.lower_frame, image=self.frame_background)
        self.frame_label.place(relwidth=1, relheight=1)


        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)


    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            # for card in self.myMagicCards:
            #     print(card.getSet())
            self.saveList(self.fileName)
            self.root.destroy()

    def get_cards(self, cardName):
        for widget in self.lower_frame.winfo_children():
            if isinstance(widget, tk.Button):
                widget.destroy()
        url = 'https://api.scryfall.com/cards/search'
        params = {'order': 'released', 'unique': 'prints', 'q': '!' + '"' + cardName + '"'}
        response = requests.get(url, params=params)
        cards = response.json()
        if cards['object']=='error':
            messagebox.showerror("Error", cards['details'])
        else:
            myList = self.make_cards(cards)
            self.make_buttons(myList)


    def make_buttons(self, cardList):
        button_x = 0
        button_y = 0
        button_list = []
        count = 0
        for card in cardList:
            url = card.getCardImage()
            image = requests.get(url)
            pic_bytes = io.BytesIO(image.content)
            pil_pic = Image.open(pic_bytes)
            tk_pic = ImageTk.PhotoImage(pil_pic)
            new_button = tk.Button(self.lower_frame, image=tk_pic,
                                   command=lambda card=card: self.addCard(card))
            new_button.image = tk_pic
            button_list.insert(count, new_button)
            button_list[count].place(x=button_x, y=button_y, width=146, height=204)
            count += 1
            button_x += 146
            if button_x == 1460:
                button_y += 204
                button_x = 0

    def make_cards(self, cards):
        card_list = []
        count = 0
        for card in cards['data']:
            card_list.append(
                MagicCard(card['name'], card['prices']['usd'], card['image_uris']['small'],
                          card['set_name'], card['rarity']))
            count += 1
        return card_list

    def addCard(self, MagicCard):
        name = MagicCard.getName()
        set = MagicCard.getSet()
        price = MagicCard.getPrice()
        rarity = MagicCard.getRarity()
        self.ws1.insert_rows(2)
        row = 2
        column=1
        values = [name, set, price, rarity]
        for value in values:
            self.ws1.cell(column=column, row=row, value=str(value))
            column+=1

    def saveList(self, saveName):
        self.wb.save(saveName)

    def run(self):
        self.root.mainloop()



